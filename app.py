
import io
import os
import re
import zipfile
import tempfile
import calendar
from datetime import datetime

import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

# PDF: pagrindinis generatorius
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors

# PDF: atsarginis (fallback) generatorius
from matplotlib.backends.backend_pdf import PdfPages
import matplotlib.pyplot as plt
import pandas as pd


# ------------------------------
# Bendri pagalbininkai
# ------------------------------
def get_current_month_end_and_name():
    """Grąžina (YYYY-MM-DD, mėnesio_pavadinimas_LT_genityvas)."""
    today = datetime.today()
    month_end_day = calendar.monthrange(today.year, today.month)[1]
    current_month_end = today.replace(day=month_end_day)
    month_names = [
        "sausio", "vasario", "kovo", "balandžio", "gegužės", "birželio",
        "liepos", "rugpjūčio", "rugsėjo", "spalio", "lapkričio", "gruodžio"
    ]
    return current_month_end.strftime("%Y-%m-%d"), month_names[today.month - 1]


def unzip_to_temp(uploaded_zip_file):
    """Išarchyvuoja į laikino katalogo šaknį ir grąžina (dir_path, tmp_handle)."""
    tmp_dir = tempfile.TemporaryDirectory()
    zip_bytes = uploaded_zip_file.read()
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        zf.extractall(tmp_dir.name)
    return tmp_dir.name, tmp_dir


def zip_tree_to_bytes(root_dir: str) -> bytes:
    """Supa©kuoja visą medį ir grąžina zip baitus atsisiuntimui."""
    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for folder_name, _, filenames in os.walk(root_dir):
            for fn in filenames:
                abs_path = os.path.join(folder_name, fn)
                arcname = os.path.relpath(abs_path, root_dir)
                zf.write(abs_path, arcname=arcname)
    mem_zip.seek(0)
    return mem_zip.read()


# ------------------------------
# PDF generatoriai
# ------------------------------
def excel_to_pdf_reportlab(xlsx_path: str, pdf_path: str):
    """
    Pagrindinis (kokybiškesnis) Excel -> PDF be MS Excel:
    - Skaito pirmą lapą.
    - Išlaiko stulpelių pločius, eilučių aukščius, paprastus rėmelius, fill, lygiavimą, bold/italic, merged cells.
    """
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        ws: Worksheet = wb.active

        # A4 ir paraštės
        page_w, page_h = A4
        margin = 12 * mm
        content_w = page_w - 2 * margin
        content_h = page_h - 2 * margin

        # Naudojamas diapazonas: print_area arba used range
        area = ws.print_area if ws.print_area else ws.calculate_dimension()
        min_col_idx, min_row_idx, max_col_idx, max_row_idx = range_boundaries(str(area))

        # Stulpelių plotis (Excel width ~ simbolių skaičius)
        col_widths = []
        for c in range(min_col_idx, max_col_idx + 1):
            letter = get_column_letter(c)
            cw = ws.column_dimensions[letter].width
            if cw is None:
                cw = 8.43  # Excel numatytasis
            # char -> px (~7 px/char), px -> pt (72/96)
            pts = cw * 7 * (72.0 / 96.0)
            col_widths.append(pts)

        # Eilučių aukštis (pt; jei None – ~15pt)
        row_heights = []
        for r in range(min_row_idx, max_row_idx + 1):
            rh = ws.row_dimensions[r].height
            if rh is None:
                rh = 15.0
            row_heights.append(float(rh))

        total_w_pts = sum(col_widths)
        total_h_pts = sum(row_heights)

        # Skalė į A4 (neišeinant už paraščių)
        scale_x = content_w / total_w_pts if total_w_pts > 0 else 1.0
        scale_y = content_h / total_h_pts if total_h_pts > 0 else 1.0
        scale = min(scale_x, scale_y, 1.0)

        origin_x = margin + (content_w - total_w_pts * scale) / 2.0
        origin_y = margin + (content_h - total_h_pts * scale) / 2.0

        # Kaupiamos sumos
        col_acc = [0.0]
        for w in col_widths:
            col_acc.append(col_acc[-1] + w * scale)
        row_acc = [0.0]
        for h in row_heights:
            row_acc.append(row_acc[-1] + h * scale)

        # Merged cells
        merged_rects = {}  # (row, col) -> (rowspan, colspan)
        for m in ws.merged_cells.ranges:
            merged_rects[(m.min_row, m.min_col)] = (m.max_row - m.min_row + 1,
                                                    m.max_col - m.min_col + 1)
        merged_members = set()
        for (sr, sc), (rs, cs) in merged_rects.items():
            for rr in range(sr, sr + rs):
                for cc in range(sc, sc + cs):
                    if not (rr == sr and cc == sc):
                        merged_members.add((rr, cc))

        def cell_xywh(r, c, rowspan=1, colspan=1):
            r0 = r - min_row_idx
            c0 = c - min_col_idx
            x = origin_x + col_acc[c0]
            # ReportLab koordinatės – apačioje; verčiam nuo viršaus
            y = origin_y + (sum(row_heights) * scale - row_acc[r0 + rowspan] + row_acc[r0])
            w = col_acc[c0 + colspan] - col_acc[c0]
            h = row_acc[r0 + rowspan] - row_acc[r0]
            return x, y, w, h

        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        c = canvas.Canvas(pdf_path, pagesize=A4)

        # 1) Užpildai + rėmeliai
        for r in range(min_row_idx, max_row_idx + 1):
            for ci in range(min_col_idx, max_col_idx + 1):
                if (r, ci) in merged_members:
                    continue
                rowspan, colspan = 1, 1
                if (r, ci) in merged_rects:
                    rowspan, colspan = merged_rects[(r, ci)]
                x, y, w, h = cell_xywh(r, ci, rowspan, colspan)

                cell = ws.cell(row=r, column=ci)

                # Užpildymas
                fill = cell.fill
                try:
                    if fill and fill.start_color and getattr(fill.start_color, "rgb", None):
                        rgb = fill.start_color.rgb  # 'FFRRGGBB'
                        if rgb and len(rgb) == 8 and rgb[2:] != "000000":
                            rr = int(rgb[2:4], 16) / 255.0
                            gg = int(rgb[4:6], 16) / 255.0
                            bb = int(rgb[6:8], 16) / 255.0
                            c.setFillColor(colors.Color(rr, gg, bb))
                            c.rect(x, y, w, h, fill=1, stroke=0)
                except Exception:
                    pass

                # Rėmeliai (paprastas atvejis)
                border = cell.border
                draw_border = any([
                    border.left and border.left.style,
                    border.right and border.right.style,
                    border.top and border.top.style,
                    border.bottom and border.bottom.style
                ])
                if draw_border:
                    c.setStrokeColor(colors.black)
                    c.setLineWidth(0.6)
                    c.rect(x, y, w, h, fill=0, stroke=1)

        # 2) Tekstas
        for r in range(min_row_idx, max_row_idx + 1):
            for ci in range(min_col_idx, max_col_idx + 1):
                if (r, ci) in merged_members:
                    continue
                rowspan, colspan = 1, 1
                if (r, ci) in merged_rects:
                    rowspan, colspan = merged_rects[(r, ci)]
                x, y, w, h = cell_xywh(r, ci, rowspan, colspan)
                cell = ws.cell(row=r, column=ci)

                val = "" if cell.value is None else str(cell.value)  # jokių NaN

                # Šriftas
                font = cell.font
                bold = bool(font and font.bold)
                italic = bool(font and font.italic)
                font_name = "Helvetica-Bold" if bold else "Helvetica"
                if italic and bold:
                    font_name = "Helvetica-BoldOblique"
                elif italic and not bold:
                    font_name = "Helvetica-Oblique"
                font_size = 9
                if font and font.sz:
                    try:
                        font_size = float(font.sz)
                    except Exception:
                        pass

                # Lygiuotė
                ha = "left"
                va = "middle"
                if cell.alignment:
                    if cell.alignment.horizontal in ("center", "centerContinuous", "distributed", "justify"):
                        ha = "center"
                    elif cell.alignment.horizontal in ("right",):
                        ha = "right"
                    if cell.alignment.vertical in ("top", "distributed", "justify"):
                        va = "top"
                    elif cell.alignment.vertical in ("bottom",):
                        va = "bottom"

                # Teksto pozicija
                pad_x = 2
                pad_y = 1
                st_x = x + pad_x if ha == "left" else (x + w / 2.0 if ha == "center" else x + w - pad_x)
                if va == "top":
                    st_y = y + h - pad_y - font_size
                elif va == "bottom":
                    st_y = y + pad_y
                else:
                    st_y = y + (h - font_size) / 2.0

                c.setFont(font_name, font_size)
                c.setFillColor(colors.black)
                if ha == "center":
                    c.drawCentredString(st_x, st_y, val)
                elif ha == "right":
                    c.drawRightString(st_x, st_y, val)
                else:
                    c.drawString(st_x, st_y, val)

        c.showPage()
        c.save()
        wb.close()
        return True, None

    except Exception as e:
        return False, str(e)


def excel_to_simple_pdf(xlsx_path: str, pdf_path: str):
    """
    Atsarginis (fallback) variantas:
    - Paimamas 1-asis lapas į pandas DataFrame.
    - Sugeneruojama paprasta lentelė į PDF (be sudėtingo formatavimo).
    """
    try:
        df = pd.read_excel(xlsx_path, sheet_name=0, header=None)
        df = df.fillna("")  # jokių NaN
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        with PdfPages(pdf_path) as pdf:
            fig, ax = plt.subplots(figsize=(11.69, 8.27))  # ~A4 landscape
            ax.axis('off')
            tbl = ax.table(cellText=df.values.astype(str), loc='center')
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(8)
            tbl.scale(1, 1.2)
            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)
        return True, None
    except Exception as e:
        return False, str(e)


# ------------------------------
# Excel failų medis: redagavimas + PDF + pervadinimas
# ------------------------------
def process_excels_in_tree(base_dir: str, log_lines: list):
    current_month_end, current_month_name = get_current_month_end_and_name()
    months = [
        "sausio", "vasario", "kovo", "balandžio", "gegužės", "birželio",
        "liepos", "rugpjūčio", "rugsėjo", "spalio", "lapkričio", "gruodžio"
    ]

    all_ok = True

    for root, _, files in os.walk(base_dir):
        rel_root = os.path.relpath(root, base_dir)
        log_lines.append(f"📁 Aplankas: {rel_root if rel_root != '.' else '/'}")
        for filename in files:
            low = filename.lower()
            if not low.endswith((".xlsx", ".xlsm")):
                continue

            file_path = os.path.join(root, filename)
            log_lines.append(f"  🔄 Failas: {os.path.join(rel_root, filename)}")

            try:
                # 1) Excel redagavimas (C5, A9)
                wb = load_workbook(file_path)
                sheet = wb.active

                # C5 – mėn. pabaigos data
                try:
                    if sheet["C5"].value is not None:
                        dt = datetime.strptime(current_month_end, "%Y-%m-%d").date()
                        sheet["C5"].value = dt
                        log_lines.append(f"    ✅ C5 -> {current_month_end}")
                except Exception as e:
                    log_lines.append(f"    ⚠️ Nepavyko atnaujinti C5: {e}")

                # A9 – mėnesio pavadinimo keitimas
                try:
                    if sheet["A9"].value is not None:
                        cell_value = str(sheet["A9"].value).strip()
                        replaced = False
                        for month in months:
                            if re.search(month, cell_value, flags=re.IGNORECASE):
                                new_val = re.sub(month, current_month_name, cell_value, flags=re.IGNORECASE)
                                sheet["A9"].value = new_val
                                replaced = True
                                log_lines.append(f"    ✅ A9 -> {new_val}")
                                break
                        if not replaced:
                            log_lines.append("    ℹ️ A9: nerastas mėnesio pavadinimas – nepakeista.")
                except Exception as e:
                    log_lines.append(f"    ⚠️ Nepavyko atnaujinti A9: {e}")

                wb.save(file_path)
                wb.close()

                # 2) Pervadinimas pagal YYYY_MM
                try:
                    year = current_month_end[:4]
                    month_num = current_month_end[5:7]
                    new_filename = re.sub(r"(\d{4})_(\d{2})", f"{year}_{month_num}", filename)
                    if new_filename != filename:
                        new_path = os.path.join(root, new_filename)
                        if os.path.exists(new_path):  # versijavimas, jei toks jau yra
                            base, ext = os.path.splitext(new_filename)
                            i = 1
                            while True:
                                candidate = os.path.join(root, f"{base}_v{i}{ext}")
                                if not os.path.exists(candidate):
                                    new_path = candidate
                                    break
                                i += 1
                        os.rename(file_path, new_path)
                        file_path = new_path
                        log_lines.append(f"    📁 Pervadinta -> {os.path.join(rel_root, os.path.basename(file_path))}")
                except Exception as e:
                    log_lines.append(f"    ⚠️ Nepavyko pervadinti: {e}")

                # 3) PDF generavimas
                try:
                    pdf_path = os.path.splitext(file_path)[0] + ".pdf"
                    if os.path.exists(pdf_path):
                        base, ext = os.path.splitext(pdf_path)
                        counter = 1
                        while True:
                            candidate = f"{base}_v{counter}{ext}"
                            if not os.path.exists(candidate):
                                pdf_path = candidate
                                break
                            counter += 1

                    ok, err = excel_to_pdf_reportlab(file_path, pdf_path)
                    if ok:
                        log_lines.append(f"    ✅ PDF (reportlab) -> {os.path.join(rel_root, os.path.basename(pdf_path))}")
                    else:
                        log_lines.append(f"    ⚠️ ReportLab nepavyko: {err} — bandau paprastą PDF...")
                        ok2, err2 = excel_to_simple_pdf(file_path, pdf_path)
                        if ok2:
                            log_lines.append(f"    ✅ PDF (simple) -> {os.path.join(rel_root, os.path.basename(pdf_path))}")
                        else:
                            all_ok = False
                            log_lines.append(f"    ❌ PDF nepavyko (abiem būdais): {err2}")
                except Exception as e:
                    all_ok = False
                    log_lines.append(f"    ❌ PDF generavimo klaida: {e}")

            except Exception as e:
                all_ok = False
                log_lines.append(f"    ❌ Apdorojimo klaida: {e}")

    return all_ok


# ------------------------------
# Streamlit UI
# ------------------------------
st.set_page_config(page_title="Aktų apdorojimas (Excel → PDF) | ZIP įkėlimas", page_icon="📄", layout="centered")

st.title("📄 Aktų apdorojimas (Streamlit Cloud)")
st.write(
    "Įkelkite **viso aplanko ZIP** (su poaplankiais). Programa atnaujins Excel failus (C5 datą, A9 mėnesį), "
    "prireikus pervadins failus `YYYY_MM` formatu, sugeneruos PDF ir grąžins visą medį kaip ZIP."
)

uploaded = st.file_uploader("Įkelkite aplanką kaip .zip", type=["zip"])

if uploaded is not None:
    with st.status("Apdorojama…", expanded=True) as status:
        logs = []
        try:
            base_dir, tmp_handle = unzip_to_temp(uploaded)
            logs.append("📦 ZIP sėkmingai išarchyvuotas.")

            all_ok = process_excels_in_tree(base_dir, logs)

            out_bytes = zip_tree_to_bytes(base_dir)
            logs.append("🧷 Paruoštas atsisiunčiamas ZIP su rezultatais.")

            status.update(label="Apdorojimas baigtas.", state="complete")

            st.text("\n".join(logs))
            st.download_button(
                label="⬇️ Parsisiųsti rezultatą (.zip)",
                data=out_bytes,
                file_name=f"apdorota_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                mime="application/zip"
            )

            if all_ok:
                st.success("🎉 Visi failai apdoroti sėkmingai!")
            else:
                st.warning("⚠️ Kai kurių failų apdoroti nepavyko. Žr. žurnalą (log).")

        except zipfile.BadZipFile:
            status.update(label="Nepavyko išarchyvuoti ZIP.", state="error")
            st.error("❌ Netinkamas ZIP failas.")
        except Exception as e:
            status.update(label="Įvyko klaida.", state="error")
            st.error(f"❌ Klaida: {e}")
else:
    st.info("👉 Pirmiausia įkelkite **.zip** su savo Excel failais.")
